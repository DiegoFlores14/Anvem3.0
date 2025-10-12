import pandas as pd
import os
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Rutas
input_folder = "./data"
output_folder = "./output"
os.makedirs(output_folder, exist_ok=True)

# Acumuladores globales
resumen_canciones = []
resumen_fuentes = []
archivos_procesados = []

def escribir_hoja(ws, df, columnas, total_royalties):
    start_row = 9
    start_col = 4  # Columna D

    # Encabezados
    for j, col in enumerate(columnas):
        ws.cell(row=start_row, column=start_col + j, value=col)

    # Datos
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row + 1):
        for j, value in enumerate(row):
            ws.cell(row=i, column=start_col + j, value=value)

    # Totales
    total_row = start_row + 1 + len(df)
    ws.cell(row=total_row, column=start_col, value="TOTAL")
    units_sum = df['UNITS'].sum() if 'UNITS' in df.columns else 0
    roy_sum = df['ROYALTIES'].sum() if 'ROYALTIES' in df.columns else 0
    ws.cell(row=total_row, column=start_col + 1, value=units_sum)
    ws.cell(row=total_row, column=start_col + 2, value=round(roy_sum, 9))

    # Net Payment en H14 / I14 (aseguramos que total_royalties siempre esté definido)
    ws.cell(row=14, column=8, value="Net Payment")
    ws.cell(row=14, column=9, value=round(total_royalties if total_royalties is not None else 0, 9))

    # Centrar contenido
    for row in ws.iter_rows(min_row=start_row, max_row=total_row,
                            min_col=start_col, max_col=start_col + len(columnas) - 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

def find_col(norm_map, alias_list):
    """
    norm_map: dict de {normalized_col_name: original_col_name}
    alias_list: lista de variantes (con espacios, mayúsculas, etc.)
    Devuelve el nombre real de la columna si lo encuentra, o None.
    """
    for alias in alias_list:
        key = ''.join(str(alias).lower().split())  # normaliza alias igual que norm_map
        if key in norm_map:
            return norm_map[key]
    return None

def procesar_archivo(file_path, output_path):
    try:
        # 1️⃣ Leer sin encabezados primero (preview)
        preview = pd.read_excel(file_path, header=None, dtype=str)

        # 2️⃣ Detectar fila del encabezado real (buscar presencia de palabras clave como substring)
        expected_keywords = ["Asset Title", "Asset Artist", "Your Earnings", "Asset Quantity", "Source"]
        header_row = None
        for i, row in preview.iterrows():
            # normalizamos cada celda a string minúscula
            row_cells = row.astype(str).str.lower().str.strip().tolist()
            # comprobamos si alguna keyword aparece como substring en alguna celda
            if any(any(k.lower() in (cell or "") for cell in row_cells) for k in expected_keywords):
                header_row = i
                break

        if header_row is None:
            print(f"⚠️ No se pudo detectar encabezado en {file_path}. Verifica el archivo.")
            return False

        # 3️⃣ Leer de nuevo con encabezado detectado
        df = pd.read_excel(file_path, header=header_row, dtype=str)  # leemos todo como str al principio
        # 4️⃣ Normalizar nombres de columnas
        df.columns = df.columns.astype(str).str.strip()

        # 5️⃣ Crear mapa de columnas normalizadas (sin espacios, todo en minúscula) -> nombre real
        norm_map = {''.join(col.lower().split()): col for col in df.columns}

        # 6️⃣ Definir aliases posibles para cada campo requerido
        aliases = {
    'assettitle': ["Asset Title", "Track", "Product Title", "Title", "AssetTitle", "release_title"],
    'assetquantity': [
        "Asset Quantity", "AssetQuantity", "Product Quantity", "ProductQuantity",
        "Quantity", "product_quantity", "asset_quantity"
    ],
    'yourearnings': ["Your Earnings", "Yourearnings", "Earnings", "Amount", "your_earnings"],
    'assetartist': [
        "Asset Artist", "Product Artist", "Artist", "AssetArtist", "release_artist"
    ],
    'source': ["Source", "Sale Store Name", "SaleStoreName", "Store", "Salestore", "sale_store_name"]
}

        # 7️⃣ Buscar columnas usando aliases
        col_title = find_col(norm_map, aliases['assettitle'])
        col_units = find_col(norm_map, aliases['assetquantity'])
        col_roy = find_col(norm_map, aliases['yourearnings'])
        col_artist = find_col(norm_map, aliases['assetartist'])
        col_source = find_col(norm_map, aliases['source'])

        # 8️⃣ Verificar columnas faltantes
        missing = []
        if col_title is None: missing.append("Asset Title (ej. 'Asset Title' / 'Track' / 'Product Title')")
        if col_units is None: missing.append("Asset Quantity (ej. 'Asset Quantity' / 'Product Quantity')")
        if col_roy is None: missing.append("Your Earnings (ej. 'Your Earnings' / 'Earnings')")
        if col_artist is None: missing.append("Asset Artist (ej. 'Asset Artist' / 'Product Artist')")
        if col_source is None: missing.append("Source (ej. 'Source' / 'Sale Store Name')")

        if missing:
            print(f"⚠️ El archivo {file_path} no contiene todas las columnas necesarias. Faltan:")
            for m in missing:
                print("   -", m)
            print("Columnas detectadas:", df.columns.tolist())
            return False

        # 9️⃣ Convertir columnas numéricas a valores correctos
        # primero extraemos las columnas reales
        df[col_units] = pd.to_numeric(df[col_units].astype(str).str.replace(",", "").str.strip(), errors='coerce').fillna(0)
        df[col_roy] = pd.to_numeric(df[col_roy].astype(str).str.replace(",", "").str.strip(), errors='coerce').fillna(0)

        # 10️⃣ Agrupar por canción (títulos + artista)
        by_song = df.groupby([col_title, col_artist], dropna=False).agg({
            col_units: 'sum',
            col_roy: 'sum'
        }).reset_index()

        # Normalizamos cadena SONG
        by_song['SONG'] = by_song[col_title].astype(str).str.strip() + " - " + by_song[col_artist].astype(str).str.strip()
        by_song = by_song.rename(columns={
            col_units: 'UNITS',
            col_roy: 'ROYALTIES'
        })[['SONG', 'UNITS', 'ROYALTIES']]

        resumen_canciones.append(by_song.copy())

        # 🔟 Agrupar por plataforma / source
        by_source = df.groupby(col_source, dropna=False).agg({
            col_units: 'sum',
            col_roy: 'sum'
        }).reset_index().rename(columns={
            col_source: 'SOURCE',
            col_units: 'UNITS',
            col_roy: 'ROYALTIES'
        })

        resumen_fuentes.append(by_source.copy())

        # 1️⃣1️⃣ Guardar archivo individual
        base, ext = os.path.splitext(output_path)
        output_path_xlsx = base + '.xlsx'
        total_royalties = by_song['ROYALTIES'].sum() if not by_song.empty else 0

        wb = Workbook()
        ws_song = wb.active
        ws_song.title = "By Song"
        ws_source = wb.create_sheet("By Source")

        escribir_hoja(ws_song, by_song, ["SONG", "UNITS", "ROYALTIES"], total_royalties)
        escribir_hoja(ws_source, by_source, ["SOURCE", "UNITS", "ROYALTIES"], total_royalties)

        wb.save(output_path_xlsx)
        print(f"✅ Guardado archivo individual: {output_path_xlsx}")
        return True

    except Exception:
        print(f"❌ Error procesando {file_path}:")
        traceback.print_exc()
        return False


# === Loop principal: procesar todos los archivos en ./data ===
for filename in os.listdir(input_folder):
    if filename.lower().endswith((".xls", ".xlsx")):
        in_file = os.path.join(input_folder, filename)

        # 📄 Cambiar "A" por "T" justo antes del número del trimestre
        nombre_salida = filename.replace("A1-", "T1-").replace("A2-", "T2-").replace("A3-", "T3-").replace("A4-", "T4-").replace("A5-", "T5-").replace("A6-", "T6-").replace("A7-", "T7-").replace("A8-", "T8-").replace("A9-", "T9-").replace("A10-", "T10-").replace("A11-", "T11-").replace("A12-", "T12-")

        out_file = os.path.join(output_folder, nombre_salida)
        ok = procesar_archivo(in_file, out_file)
        if ok:
            archivos_procesados.append(nombre_salida)


# === Crear resumen global ===
if resumen_canciones and resumen_fuentes:
    resumen_path = os.path.join(output_folder, "resumen_por_artista.xlsx")
    base, ext = os.path.splitext(resumen_path)
    resumen_path = base + '.xlsx'

    wb = Workbook()

    # Canciones
    df_canciones = pd.concat(resumen_canciones, ignore_index=True)
    todas_canciones = df_canciones.groupby("SONG", as_index=False).agg({
        "UNITS": "sum",
        "ROYALTIES": "sum"
    }).sort_values("ROYALTIES", ascending=False)

    total_royalties_songs = todas_canciones['ROYALTIES'].sum() if not todas_canciones.empty else 0

    ws_song = wb.active
    ws_song.title = "By Song"
    escribir_hoja(ws_song, todas_canciones, ["SONG", "UNITS", "ROYALTIES"], total_royalties_songs)

    # Plataformas
    df_fuentes = pd.concat(resumen_fuentes, ignore_index=True)
    todas_fuentes = df_fuentes.groupby("SOURCE", as_index=False).agg({
        "UNITS": "sum",
        "ROYALTIES": "sum"
    }).sort_values("ROYALTIES", ascending=False)

    total_royalties_sources = todas_fuentes['ROYALTIES'].sum() if not todas_fuentes.empty else 0

    ws_source = wb.create_sheet("By Source")
    escribir_hoja(ws_source, todas_fuentes, ["SOURCE", "UNITS", "ROYALTIES"], total_royalties_sources)

    wb.save(resumen_path)
    print(f"✅ Guardado resumen global: {resumen_path}")

# === Resumen en consola ===
print("\n--- Resumen de ejecución ---")
if archivos_procesados:
    print("Archivos individuales procesados:")
    for a in archivos_procesados:
        print(" -", a)
else:
    print("No se procesó ningún archivo individual.")

print("Archivos en carpeta output:")
for f in os.listdir(output_folder):
    print(" -", f)

print("\nProceso finalizado.")
